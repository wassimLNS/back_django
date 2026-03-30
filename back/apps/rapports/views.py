from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncDay, TruncMonth
from django.utils import timezone
from datetime import timedelta

from apps.tickets.models import Ticket, TypeService
from apps.users.models import Utilisateur
from apps.users.permissions import EstAdmin


# ============================================================
# STATISTIQUES GÉNÉRALES DU CENTRE (admin)
# ============================================================
class StatsGeneralesView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        centre = request.user.centre
        if not centre:
            return Response({'error': 'Pas de centre associé'})

        tickets = Ticket.objects.filter(centre=centre)

        # Stats globales
        total           = tickets.count()
        ouverts         = tickets.filter(statut='ouvert').count()
        en_cours        = tickets.filter(statut='en_cours').count()
        resolus         = tickets.filter(statut='resolu').count()
        fermes          = tickets.filter(statut='ferme').count()
        escalades       = tickets.filter(statut__in=['escalade_technique', 'escalade_annexe']).count()
        rejetes         = tickets.filter(statut='rejete').count()

        # Satisfaction moyenne
        satisfaction = tickets.filter(
            satisfaction_client__isnull=False
        ).aggregate(moy=Avg('satisfaction_client'))['moy']

        # Tickets par priorité
        par_priorite = dict(tickets.values('priorite').annotate(count=Count('id')).values_list('priorite', 'count'))

        # Tickets par type de service
        par_type = list(tickets.values(
            libelle=F('type_service__libelle')
        ).annotate(count=Count('id')).order_by('-count')[:5])

        # Tickets des 7 derniers jours
        sept_jours = timezone.now() - timedelta(days=7)
        par_jour = list(tickets.filter(
            created_at__gte=sept_jours
        ).annotate(
            jour=TruncDay('created_at')
        ).values('jour').annotate(count=Count('id')).order_by('jour'))

        # Tickets par mois (6 derniers mois)
        six_mois = timezone.now() - timedelta(days=180)
        par_mois = list(tickets.filter(
            created_at__gte=six_mois
        ).annotate(
            mois=TruncMonth('created_at')
        ).values('mois').annotate(count=Count('id')).order_by('mois'))

        return Response({
            'total':            total,
            'ouverts':          ouverts,
            'en_cours':         en_cours,
            'resolus':          resolus,
            'fermes':           fermes,
            'escalades':        escalades,
            'rejetes':          rejetes,
            'satisfaction_moy': round(satisfaction, 2) if satisfaction else None,
            'par_priorite':     par_priorite,
            'par_type_service': par_type,
            'par_jour':         par_jour,
            'par_mois':         par_mois,
        })


# ============================================================
# PERFORMANCES DES AGENTS (admin)
# ============================================================
class PerformancesAgentsView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        centre = request.user.centre
        if not centre:
            return Response({'error': 'Pas de centre associé'})

        agents = Utilisateur.objects.filter(
            centre=centre,
            role__in=['agent', 'agent_technique', 'agent_annexe'],
            actif=True
        )

        performances = []
        for agent in agents:
            tickets = Ticket.objects.filter(agent=agent)
            total   = tickets.count()

            resolus  = tickets.filter(statut__in=['resolu', 'ferme']).count()
            escalades = tickets.filter(statut__in=['escalade_technique', 'escalade_annexe']).count()
            actifs   = tickets.filter(statut__in=['ouvert', 'en_cours']).count()

            # Temps moyen de résolution en minutes
            temps_resolution = tickets.filter(
                resolu_a__isnull=False,
                pris_en_charge_a__isnull=False
            ).aggregate(
                moy=Avg(
                    ExpressionWrapper(
                        F('resolu_a') - F('pris_en_charge_a'),
                        output_field=DurationField()
                    )
                )
            )['moy']

            moy_minutes = None
            if temps_resolution:
                moy_minutes = round(temps_resolution.total_seconds() / 60, 0)

            # Satisfaction moyenne
            satisfaction = tickets.filter(
                satisfaction_client__isnull=False
            ).aggregate(moy=Avg('satisfaction_client'))['moy']

            performances.append({
                'agent_id':          str(agent.id),
                'nom':               agent.nom,
                'prenom':            agent.prenom,
                'role':              agent.role,
                'total_tickets':     total,
                'tickets_actifs':    actifs,
                'tickets_resolus':   resolus,
                'tickets_escalades': escalades,
                'taux_resolution':   round(resolus / total * 100, 1) if total > 0 else 0,
                'taux_escalade':     round(escalades / total * 100, 1) if total > 0 else 0,
                'moy_resolution_min': moy_minutes,
                'satisfaction_moy':  round(satisfaction, 2) if satisfaction else None,
                'derniere_connexion': agent.derniere_connexion,
            })

        # Trier par taux de résolution décroissant
        performances.sort(key=lambda x: x['taux_resolution'], reverse=True)
        return Response(performances)


# ============================================================
# EXPORT PDF
# ============================================================
class ExportPDFView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from django.http import HttpResponse
        import io

        centre = request.user.centre
        if not centre:
            return Response({'error': 'Pas de centre associé'})

        buffer   = io.BytesIO()
        doc      = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles   = getSampleStyleSheet()

        # Titre
        elements.append(Paragraph(f"Rapport — {centre.nom}", styles['Title']))
        elements.append(Paragraph(f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Stats globales
        tickets = Ticket.objects.filter(centre=centre)
        elements.append(Paragraph("Statistiques Générales", styles['Heading2']))
        data = [
            ['Indicateur', 'Valeur'],
            ['Total tickets',   tickets.count()],
            ['Ouverts',         tickets.filter(statut='ouvert').count()],
            ['En cours',        tickets.filter(statut='en_cours').count()],
            ['Résolus',         tickets.filter(statut='resolu').count()],
            ['Fermés',          tickets.filter(statut='ferme').count()],
            ['Escaladés',       tickets.filter(statut__in=['escalade_technique', 'escalade_annexe']).count()],
        ]
        table = Table(data, colWidths=[300, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#006994')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Performances agents
        elements.append(Paragraph("Performances des Agents", styles['Heading2']))
        agents = Utilisateur.objects.filter(centre=centre, role='agent', actif=True)
        data2  = [['Agent', 'Total', 'Résolus', 'Escaladés', 'Taux résolution']]
        for agent in agents:
            t = Ticket.objects.filter(agent=agent)
            total    = t.count()
            resolus  = t.filter(statut__in=['resolu', 'ferme']).count()
            escalades = t.filter(statut__in=['escalade_technique', 'escalade_annexe']).count()
            taux     = f"{round(resolus / total * 100, 1)}%" if total > 0 else "0%"
            data2.append([f"{agent.prenom} {agent.nom}", total, resolus, escalades, taux])

        table2 = Table(data2, colWidths=[150, 70, 70, 80, 100])
        table2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#006994')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table2)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{centre.code}_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response


# ============================================================
# EXPORT EXCEL
# ============================================================
class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io

        centre = request.user.centre
        if not centre:
            return Response({'error': 'Pas de centre associé'})

        wb = Workbook()

        # ---- Feuille 1 : Tickets ----
        ws1 = wb.active
        ws1.title = "Tickets"
        headers = ['N° Ticket', 'Titre', 'Client', 'Agent', 'Type', 'Statut', 'Priorité', 'Créé le', 'Résolu le']
        ws1.append(headers)

        # Style entête
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(fill_type="solid", fgColor="006994")
            cell.alignment = Alignment(horizontal='center')

        tickets = Ticket.objects.filter(centre=centre).select_related('client', 'agent', 'type_service')
        for t in tickets:
            ws1.append([
                t.numero_ticket,
                t.titre,
                f"{t.client.prenom} {t.client.nom}",
                f"{t.agent.prenom} {t.agent.nom}" if t.agent else "Non attribué",
                t.type_service.libelle,
                t.statut,
                t.priorite,
                t.created_at.strftime('%d/%m/%Y %H:%M') if t.created_at else '',
                t.resolu_a.strftime('%d/%m/%Y %H:%M') if t.resolu_a else '',
            ])

        # ---- Feuille 2 : Performances ----
        ws2 = wb.create_sheet("Performances Agents")
        headers2 = ['Agent', 'Rôle', 'Total', 'Résolus', 'Escaladés', 'Taux résolution', 'Satisfaction moy']
        ws2.append(headers2)

        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(fill_type="solid", fgColor="006994")
            cell.alignment = Alignment(horizontal='center')

        agents = Utilisateur.objects.filter(centre=centre, role__in=['agent', 'agent_technique', 'agent_annexe'], actif=True)
        for agent in agents:
            t        = Ticket.objects.filter(agent=agent)
            total    = t.count()
            resolus  = t.filter(statut__in=['resolu', 'ferme']).count()
            escalades = t.filter(statut__in=['escalade_technique', 'escalade_annexe']).count()
            taux     = f"{round(resolus / total * 100, 1)}%" if total > 0 else "0%"
            satisfaction = t.filter(satisfaction_client__isnull=False).aggregate(moy=Avg('satisfaction_client'))['moy']
            ws2.append([
                f"{agent.prenom} {agent.nom}",
                agent.role,
                total,
                resolus,
                escalades,
                taux,
                round(satisfaction, 2) if satisfaction else '',
            ])

        # Ajuster largeur colonnes
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{centre.code}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response